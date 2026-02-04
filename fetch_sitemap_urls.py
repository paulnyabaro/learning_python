import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

def setup_driver():
    """Set up Chrome driver with options."""
    chrome_options = Options()
    chrome_options.add_argument('--headless')  # Run in background
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def fetch_sitemap_urls(sitemap_url, driver):
    """Fetch and parse sitemap XML to extract URLs using Selenium."""
    try:
        print(f"  Fetching {sitemap_url}...")
        driver.get(sitemap_url)
        time.sleep(5)  # Wait longer for page to load
        
        # Get page source (XML content)
        page_source = driver.page_source
        
        # Debug: Print first 500 chars to see what we got
        print(f"  Content preview: {page_source[:500]}...")
        
        urls = []
        
        # Method 1: Regex extraction (most robust for mixed content)
        import re
        # Look for <loc>URL</loc> pattern, handling potential whitespace
        # User specified structure: <loc>https://...</loc>
        matches = re.findall(r'<loc>\s*(https?://[^<]+?)\s*</loc>', page_source)
        if matches:
            print(f"  Extracted {len(matches)} URLs using regex.")
            urls.extend(matches)
            
        # Method 2: Fallback to finding 'text' in specific elements if regex fails
        if not urls:
             # Sometimes Chrome wraps XML in HTML with specific structure
             try:
                 # If it's pure text presented
                 elements = driver.find_elements(By.TAG_NAME, "loc")
                 for el in elements:
                     urls.append(el.text)
                 if urls:
                     print(f"  Extracted {len(urls)} URLs using Selenium elements.")
             except Exception as e:
                 print(f"  Selenium element extraction failed: {e}")

        return urls
    except Exception as e:
        print(f"  Error fetching {sitemap_url}: {e}")
        return []

def save_to_excel(all_urls, filename='sitemap_urls.xlsx'):
    """Save URLs to Excel file."""
    # Create workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sitemap URLs"
    
    # Add header
    ws['A1'] = 'No.'
    ws['B1'] = 'URL'
    ws['C1'] = 'Source Sitemap'
    
    # Style header
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='center')
    
    # Add URLs
    row = 2
    for url_data in all_urls:
        ws[f'A{row}'] = row - 1
        ws[f'B{row}'] = url_data['url']
        ws[f'C{row}'] = url_data['source']
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 40
    
    # Save workbook
    wb.save(filename)
    print(f"\n✓ Successfully saved {row-2} URLs to {filename}")

def main():
    # Sitemap URLs
    sitemaps = [
        'https://legalwritingexperts.com/sitemap-posts-1.xml'
    ]
    
    all_urls = []
    
    print("Setting up browser...")
    driver = setup_driver()
    
    try:
        print("Fetching URLs from sitemaps...\n")
        
        # Fetch URLs from each sitemap
        for sitemap_url in sitemaps:
            print(f"Processing: {sitemap_url}")
            urls = fetch_sitemap_urls(sitemap_url, driver)
            print(f"  Found {len(urls)} URLs")
            
            # Add to collection with source information
            for url in urls:
                all_urls.append({
                    'url': url,
                    'source': sitemap_url
                })
        
        # Save to Excel
        if all_urls:
            save_to_excel(all_urls)
            print(f"\nTotal URLs collected: {len(all_urls)}")
        else:
            print("\nNo URLs found!")
    
    finally:
        driver.quit()
        print("\nBrowser closed.")

if __name__ == "__main__":
    main()
